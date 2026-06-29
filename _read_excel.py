import openpyxl
wb = openpyxl.load_workbook('/Users/shuo/product/product_consistency_summary.xlsx')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'Sheet: {s}')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Merged: {ws.merged_cells.ranges}')
    for row in ws.iter_rows(max_row=30):
        cells = [(c.coordinate, c.value) for c in row if c.value is not None]
        if cells:
            print(cells)
    print('---')
