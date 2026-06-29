"""
xlsx 数据集加载器
- 把 xlsx 中嵌入的图片提取为 base64 data URL（VLM API 可直接消费）
- 自动按行聚合：主体图 / logo / 生成图 / 赠品图 / 是否0分 等字段
"""
import base64
import openpyxl
from typing import List, Dict, Any


def load_dataset(xlsx_path: str) -> List[Dict[str, Any]]:
    """
    加载 xlsx，每行返回一个字典：
    {
      "row": 2,
      "主体图": "data:image/jpeg;base64,...",
      "生成图": "data:image/jpeg;base64,...",
      "logo":   "data:image/jpeg;base64,...",   # 可能不存在
      "是否0分": ..., "0分原因": ..., ...
    }
    只保留至少包含「生成图」的有效行。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 表头（列名 → 0-based 列索引）
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    name_to_col = {name: i for i, name in enumerate(headers) if name}

    # 收集嵌入图片：行号(1-based) → {列索引(0-based): data_url}
    row_images: Dict[int, Dict[int, str]] = {}
    for img in ws._images:
        f = img.anchor._from
        row_1based = f.row + 1
        col_0based = f.col
        fmt = (img.format or "jpeg").lower()
        try:
            data = img._data()
        except Exception:
            continue
        b64 = base64.b64encode(data).decode()
        data_url = f"data:image/{fmt};base64,{b64}"
        row_images.setdefault(row_1based, {})[col_0based] = data_url

    # 组装记录
    records: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        rec: Dict[str, Any] = {"row": r}
        # 文本字段
        for name, col in name_to_col.items():
            val = ws.cell(r, col + 1).value
            if val is not None:
                rec[name] = val
        # 图片字段（覆盖同名文本字段，因为图片单元格本身可能为空文本）
        for name, col in name_to_col.items():
            if col in row_images.get(r, {}):
                rec[name] = row_images[r][col]
        # 必须有「生成图」
        if isinstance(rec.get("生成图"), str) and rec["生成图"].startswith("data:image/"):
            records.append(rec)
    return records


if __name__ == "__main__":
    recs = load_dataset("data/dataset_without_gift.xlsx")
    print(f"有效记录数: {len(recs)}")
    if recs:
        sample = recs[0]
        for k, v in sample.items():
            preview = v[:60] + "..." if isinstance(v, str) and len(v) > 60 else v
            print(f"  {k}: {preview}")
